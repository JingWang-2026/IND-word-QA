from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

from app.models.document import Document
from app.models.issue import QAIssue

EXPORT_COLUMNS = [
    "Issue ID",
    "Severity",
    "Category",
    "Rule ID",
    "Document Name",
    "Section Number",
    "Section Title",
    "Location",
    "Source Text",
    "Description",
    "Suggestion",
    "Status",
    "Reviewer Comment",
    "Created At",
]


class ExportService:
    def issues_xlsx(self, session: Session, issues: list[QAIssue]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "QA Issues"
        sheet.append(EXPORT_COLUMNS)
        self._style_header(sheet)

        document_names = self._document_names(session, issues)
        for issue in issues:
            location = issue.location_json or ""
            section_number, section_title = self._section_fields(location)
            sheet.append(
                [
                    issue.id,
                    issue.severity,
                    issue.category,
                    issue.rule_id or "",
                    document_names.get(issue.document_id or "", ""),
                    section_number,
                    section_title,
                    location,
                    issue.source_text or "",
                    issue.description,
                    issue.suggestion or "",
                    issue.status,
                    issue.reviewer_comment or "",
                    issue.created_at.isoformat(),
                ]
            )

        sheet.freeze_panes = "A2"
        self._fit_columns(sheet)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _document_names(self, session: Session, issues: list[QAIssue]) -> dict[str, str]:
        ids = sorted({issue.document_id for issue in issues if issue.document_id})
        if not ids:
            return {}
        documents = session.exec(select(Document).where(Document.id.in_(ids))).all()
        return {document.id: document.filename for document in documents}

    def _style_header(self, sheet: Worksheet) -> None:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    def _fit_columns(self, sheet: Worksheet) -> None:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_length + 2, 12), 60)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _section_fields(self, location_json: str) -> tuple[str, str]:
        if '"section_number"' not in location_json and '"section_title"' not in location_json:
            return "", ""
        import json

        try:
            location = json.loads(location_json)
        except json.JSONDecodeError:
            return "", ""
        return str(location.get("section_number") or ""), str(location.get("section_title") or "")
